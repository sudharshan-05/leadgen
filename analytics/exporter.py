"""
analytics/exporter.py — Professional Excel Export Engine.
Generates multi-sheet Apollo-style Excel reports and segments leads into platform-specific files.
"""
from __future__ import annotations

import os
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import config

# Styles definitions
FONT_FAMILY = "Calibri"
FONT_SIZE = 11

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
HEADER_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE, bold=True, color="FFFFFF")

HOT_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid") # Green
HOT_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE, color="FFFFFF")

WARM_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid") # Amber
WARM_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE, color="000000")

COLD_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") # Red
COLD_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE, color="FFFFFF")

SKIP_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid") # Gray
SKIP_FONT = Font(name=FONT_FAMILY, size=FONT_SIZE, color="000000")

THIN_BORDER = Border(
    left=Side(style='thin', color='D3D3D3'),
    right=Side(style='thin', color='D3D3D3'),
    top=Side(style='thin', color='D3D3D3'),
    bottom=Side(style='thin', color='D3D3D3')
)

COLUMNS_SCHEMA = [
    # Identity & Segment
    "Rank", "Business Name", "Industry Detected", "Source",
    # Activity Level (Layer 1)
    "Business Active", "Activity Score",
    # Contact
    "Phone", "Email", "Email Confidence",
    # Online Presence (Layer 2 & 3)
    "Website (Existing)", "Website Status", "Status Confidence", "Is Outdated", "Copyright Year",
    "Instagram URL", "Facebook URL", "LinkedIn URL",
    # Location
    "Address", "City",
    # Reputation
    "Rating", "Reviews",
    # Technical Check (Layer 2)
    "HTTPS Enabled", "SSL Valid", "Mobile Ready", "Load Time (s)", "Page Speed Score", "Website Score", "CMS Detected",
    "Has Chatbot", "Has Booking", "Has WhatsApp", "Has Contact Form", "Has Sitemap", "Has Pricing Page",
    "Detected Technologies",
    # Business Intelligence (Layer 4)
    "Company Size", "Budget Level", "Company Summary",
    # Pain Points & Pitching (Layer 5, 6, 7)
    "Pain Points", "Recommended Services", "Outreach Angle / Pitch",
    # Scoring & Qualification (Layer 8)
    "Opportunity Score", "Tier",
    # Reference
    "Google Maps URL", "Place ID"
]

def format_cell_value(val) -> str:
    """Format Boolean fields into Yes/No."""
    if val is True or val == 1 or val == "Yes":
        return "Yes"
    if val is False or val == 0 or val == "No":
        return "No"
    return str(val) if val is not None else ""

def apply_auto_width_and_borders(ws: openpyxl.worksheet.worksheet.Worksheet):
    """Adjust column widths dynamically, add cell borders, and set alignments."""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            cell.border = THIN_BORDER
            if cell.font is None or cell.font.name != FONT_FAMILY:
                cell.font = Font(name=FONT_FAMILY, size=FONT_SIZE)
            val_str = str(cell.value or "")
            if len(val_str) > max_len:
                max_len = len(val_str)

        ws.column_dimensions[col_letter].width = max(10, min(max_len + 3, 40))

    if ws.max_column >= 41:
        ws.column_dimensions[get_column_letter(38)].width = 35
        ws.column_dimensions[get_column_letter(39)].width = 35
        ws.column_dimensions[get_column_letter(40)].width = 35
        ws.column_dimensions[get_column_letter(41)].width = 50

def write_headers(ws: openpyxl.worksheet.worksheet.Worksheet, columns: list[str]):
    """Write standard headers onto a worksheet."""
    ws.append(columns)
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

def add_lead_row(ws: openpyxl.worksheet.worksheet.Worksheet, rank: int, lead: dict, extra_cols: list | None = None):
    """Add a lead record row onto a sheet and style it according to its tier."""
    score = lead.get("opportunity_score", 0)
    tier = lead.get("tier", "SKIP")

    row_fill, row_font = None, None
    if tier == "HOT":
        row_fill, row_font = HOT_FILL, HOT_FONT
    elif tier == "WARM":
        row_fill, row_font = WARM_FILL, WARM_FONT
    elif tier == "COLD":
        row_fill, row_font = COLD_FILL, COLD_FONT
    else:
        row_fill, row_font = SKIP_FILL, SKIP_FONT

    is_live_status = lead.get("website_status_label", "no_website")
    if is_live_status == "no_website":
        is_live_status = "❌ No Website"
    elif is_live_status == "outdated":
        is_live_status = "⚠️ Outdated"
    elif is_live_status == "average":
        is_live_status = "⚖️ Average"
    elif is_live_status == "modern":
        is_live_status = "✅ Modern"

    web = lead.get("website") or ""
    src = lead.get("source", "")
    insta_url = web if "instagram.com" in web.lower() or src == "instagram" else ""
    fb_url    = web if "facebook.com"  in web.lower() or src == "facebook"  else ""
    li_url    = web if "linkedin.com"  in web.lower() or src == "linkedin"  else ""
    main_website = web if not (insta_url or fb_url or li_url) else ""

    load_t = lead.get("load_time")
    load_str = f"{float(load_t):.2f}" if load_t else ""

    cop_year = lead.get("copyright_year", 0)
    cop_str = str(cop_year) if cop_year and int(cop_year) > 2000 else ""

    import json as _json
    try:
        tech_list = ", ".join(_json.loads(lead.get("detected_technologies") or "[]"))
    except Exception:
        tech_list = lead.get("detected_technologies") or ""

    try:
        pains = ", ".join(_json.loads(lead.get("pain_points") or "[]"))
    except Exception:
        pains = lead.get("pain_points") or ""

    try:
        svcs = ", ".join(_json.loads(lead.get("recommended_services") or "[]"))
    except Exception:
        svcs = lead.get("recommended_services") or ""

    row_data = [
        rank,
        lead.get("business_name", ""),
        lead.get("industry_detected", lead.get("category", "")),
        lead.get("source", ""),
        
        format_cell_value(lead.get("business_active")),
        lead.get("activity_score", ""),
        
        lead.get("phone", ""),
        lead.get("email", ""),
        lead.get("email_confidence", "none"),
        
        main_website,
        is_live_status,
        lead.get("website_status_confidence", ""),
        format_cell_value(lead.get("is_outdated")),
        cop_str,
        insta_url,
        fb_url,
        li_url,
        
        lead.get("address", ""),
        lead.get("city", ""),
        
        lead.get("rating", 0.0),
        lead.get("review_count", 0),
        
        format_cell_value(lead.get("has_https")),
        format_cell_value(lead.get("ssl_valid")),
        format_cell_value(lead.get("is_mobile_ready")),
        load_str,
        lead.get("page_speed_score", ""),
        lead.get("website_score", ""),
        lead.get("cms_detected", ""),
        format_cell_value(lead.get("has_chatbot")),
        format_cell_value(lead.get("has_booking")),
        format_cell_value(lead.get("has_whatsapp")),
        format_cell_value(lead.get("has_contact_form")),
        format_cell_value(lead.get("has_sitemap")),
        format_cell_value(lead.get("has_pricing_page")),
        tech_list,
        
        lead.get("company_size", ""),
        lead.get("budget_level", ""),
        lead.get("company_summary", ""),
        
        pains,
        svcs,
        lead.get("outreach_angle", lead.get("recommended_pitch", "")),
        
        score,
        tier,
        
        lead.get("google_maps_url", ""),
        lead.get("place_id", ""),
    ]

    if extra_cols:
        row_data.extend(extra_cols)

    ws.append(row_data)
    curr_row = ws.max_row

    CENTER_COLS = {1, 5, 6, 11, 12, 13, 14, 20, 21, 22, 23, 24, 26, 27, 29, 30, 31, 32, 33, 34, 42, 43}
    for col_idx in range(1, len(row_data) + 1):
        cell = ws.cell(row=curr_row, column=col_idx)
        cell.fill = row_fill
        cell.font = row_font
        cell.alignment = Alignment(
            horizontal="center" if col_idx in CENTER_COLS else "left",
            wrap_text=(col_idx in (38, 39, 40, 41))
        )

def export_to_excel(leads: list[dict], industry: str, location: str) -> str:
    """
    Export all leads into a professional multi-sheet Excel spreadsheet.
    Generates segment files for Instagram, LinkedIn, and Facebook leads.
    """
    now = datetime.now()
    date_str = now.strftime("%Y-%m-%d")
    time_str = now.strftime("%H%M")
    
    ind_clean = "".join(x for x in industry.title() if x.isalnum())
    loc_clean = "".join(x for x in location.title() if x.isalnum())
    
    filename = f"{ind_clean}_{loc_clean}_{date_str}_{time_str}.xlsx"
    filepath = os.path.join(str(config.OUTPUTS_FOLDER), filename)
    
    _NON_BIZ_KW = ["house", "home", "flat", "kuppam", "illam", "veedu", "room", "layout", "colony"]

    def _is_quality_lead(lead: dict) -> bool:
        name = (lead.get("business_name") or "").strip()
        if not name or len(name) < 4:
            return False
        if not any(c.isalpha() for c in name):
            return False
        if any(kw in name.lower() for kw in _NON_BIZ_KW):
            return False
        has_phone = bool(lead.get("phone") and str(lead["phone"]).strip())
        has_maps  = bool(lead.get("google_maps_url") and str(lead["google_maps_url"]).strip())
        has_place = bool(lead.get("place_id") and str(lead["place_id"]).strip())
        has_email = bool(lead.get("email") and str(lead["email"]).strip())
        has_web   = bool(lead.get("website") and str(lead["website"]).strip())
        return has_phone or has_maps or has_place or has_email or has_web

    before_export = len(leads)
    leads = [l for l in leads if _is_quality_lead(l)]
    if len(leads) < before_export:
        print(f"[Exporter] Dropped {before_export - len(leads)} junk leads before Excel export.")

    sorted_leads = sorted(leads, key=lambda x: x.get("opportunity_score", 0), reverse=True)
    
    wb = openpyxl.Workbook()
    
    # Summary Sheet
    ws_dash = wb.active
    ws_dash.title = "Executive Summary"
    ws_dash.views.sheetView[0].showGridLines = True
    
    ws_dash.merge_cells("A1:D1")
    title_cell = ws_dash["A1"]
    title_cell.value = "LEAD INTELLIGENCE REPORT"
    title_cell.font = Font(name=FONT_FAMILY, size=16, bold=True, color="FFFFFF")
    title_cell.fill = HEADER_FILL
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_dash.row_dimensions[1].height = 40
    
    tot_leads = len(sorted_leads)
    hot_count = sum(1 for l in sorted_leads if l.get("tier") == "HOT")
    warm_count = sum(1 for l in sorted_leads if l.get("tier") == "WARM")
    email_count = sum(1 for l in sorted_leads if l.get("email_confidence") == "found")
    no_web_count = sum(1 for l in sorted_leads if not l.get("website"))
    outdated_count = sum(1 for l in sorted_leads if l.get("is_outdated", 0) == 1)
    no_chat_count = sum(1 for l in sorted_leads if l.get("has_chatbot", 1) == 0)
    
    avg_score = sum(l.get("opportunity_score", 0) for l in sorted_leads) / tot_leads if tot_leads > 0 else 0
    
    source_counts = {}
    for l in sorted_leads:
        src = l.get("source", "unknown")
        source_counts[src] = source_counts.get(src, 0) + 1
        
    dashboard_data = [
        ("Total Leads Discovered", tot_leads),
        ("HOT Opportunity Leads (>=80)", hot_count),
        ("WARM Opportunity Leads (60-79)", warm_count),
        ("Emails Successfully Found", email_count),
        ("No Website Opportunities", no_web_count),
        ("Outdated Website Opportunities", outdated_count),
        ("No Chatbot Opportunities", no_chat_count),
        ("Average Opportunity Score", f"{avg_score:.1f}"),
        ("Generation Timestamp", now.strftime("%Y-%m-%d %H:%M:%S"))
    ]
    
    ws_dash.append([])
    ws_dash.append(["Metrics Summary", "Value"])
    ws_dash.cell(row=3, column=1).font = Font(name=FONT_FAMILY, size=12, bold=True)
    ws_dash.cell(row=3, column=2).font = Font(name=FONT_FAMILY, size=12, bold=True)
    
    for metric, val in dashboard_data:
        ws_dash.append([metric, val])
        
    ws_dash.append([])
    ws_dash.append(["Scraper Source", "Leads Scraped"])
    ws_dash.cell(row=ws_dash.max_row, column=1).font = Font(name=FONT_FAMILY, size=12, bold=True)
    ws_dash.cell(row=ws_dash.max_row, column=2).font = Font(name=FONT_FAMILY, size=12, bold=True)
    
    for src, count in source_counts.items():
        ws_dash.append([src.replace("_", " ").title(), count])
        
    apply_auto_width_and_borders(ws_dash)
    
    # All Leads
    ws_all = wb.create_sheet(title="All Leads")
    ws_all.views.sheetView[0].showGridLines = True
    ws_all.freeze_panes = "A2"
    
    write_headers(ws_all, COLUMNS_SCHEMA)
    for idx, lead in enumerate(sorted_leads, start=1):
        add_lead_row(ws_all, idx, lead)
        
    ws_all.auto_filter.ref = f"A1:{get_column_letter(ws_all.max_column)}{ws_all.max_row}"
    apply_auto_width_and_borders(ws_all)
    
    # HOT Leads
    ws_hot = wb.create_sheet(title="HOT Leads Only")
    ws_hot.views.sheetView[0].showGridLines = True
    ws_hot.freeze_panes = "A2"
    
    write_headers(ws_hot, COLUMNS_SCHEMA)
    hot_idx = 1
    for lead in sorted_leads:
        if lead.get("tier") == "HOT":
            add_lead_row(ws_hot, hot_idx, lead)
            hot_idx += 1
            
    apply_auto_width_and_borders(ws_hot)
    
    # No Website
    ws_noweb = wb.create_sheet(title="No Website")
    ws_noweb.views.sheetView[0].showGridLines = True
    ws_noweb.freeze_panes = "A2"
    
    noweb_headers = COLUMNS_SCHEMA + ["Estimated Project Value"]
    write_headers(ws_noweb, noweb_headers)
    noweb_idx = 1
    for lead in sorted_leads:
        if not lead.get("website"):
            add_lead_row(ws_noweb, noweb_idx, lead, extra_cols=["₹25,000 to ₹75,000 project"])
            noweb_idx += 1
            
    apply_auto_width_and_borders(ws_noweb)
    
    # Outdated
    ws_old = wb.create_sheet(title="Outdated Website")
    ws_old.views.sheetView[0].showGridLines = True
    ws_old.freeze_panes = "A2"
    
    old_headers = COLUMNS_SCHEMA + ["Last Updated"]
    write_headers(ws_old, old_headers)
    old_idx = 1
    for lead in sorted_leads:
        if lead.get("is_outdated", 0) == 1:
            cop_year = lead.get("copyright_year", 0)
            add_lead_row(ws_old, old_idx, lead, extra_cols=[str(cop_year) if cop_year > 0 else "Unknown"])
            old_idx += 1
            
    apply_auto_width_and_borders(ws_old)
    
    # Quick Wins
    ws_wins = wb.create_sheet(title="Quick Wins")
    ws_wins.views.sheetView[0].showGridLines = True
    ws_wins.freeze_panes = "A2"
    
    write_headers(ws_wins, COLUMNS_SCHEMA)
    wins_idx = 1
    for lead in sorted_leads:
        if lead.get("email_confidence") == "found" and lead.get("opportunity_score", 0) >= 60:
            add_lead_row(ws_wins, wins_idx, lead)
            wins_idx += 1
            
    apply_auto_width_and_borders(ws_wins)
    
    wb.save(filepath)
    print(f"[Excel Exporter] Master report created: {filepath}")
    
    # Platform-specific segments
    for platform in ["instagram", "linkedin", "facebook"]:
        plat_leads = [
            l for l in sorted_leads
            if l.get("source") == platform or platform in str(l.get("website", "")).lower()
        ]
        
        if plat_leads:
            plat_fn = f"{platform.title()}_Leads_{date_str}_{time_str}.xlsx"
            plat_fp = os.path.join(str(config.OUTPUTS_FOLDER), plat_fn)
            
            plat_wb = openpyxl.Workbook()
            plat_ws = plat_wb.active
            plat_ws.title = f"{platform.title()} Leads"
            plat_ws.views.sheetView[0].showGridLines = True
            plat_ws.freeze_panes = "A2"
            
            write_headers(plat_ws, COLUMNS_SCHEMA)
            for idx, p_lead in enumerate(plat_leads, start=1):
                lead_copy = {**p_lead, "source": platform}
                add_lead_row(plat_ws, idx, lead_copy)
                
            plat_ws.auto_filter.ref = f"A1:{get_column_letter(plat_ws.max_column)}{plat_ws.max_row}"
            apply_auto_width_and_borders(plat_ws)
            plat_wb.save(plat_fp)
            print(f"[Excel Exporter] Platform segment report created: {plat_fp}")
            
    return filepath
